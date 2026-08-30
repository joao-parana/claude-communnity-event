"""Pré-agrega as bases do desafio no que o painel mostra.
Roda offline, uma vez; o painel lê só o JSON resultante."""
import json, os, collections, math, re, sys
import openpyxl
import duckdb

_B = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "dadoscreche", "Bases IC_ ClassificadoseFila")
c = duckdb.connect()
c.execute(f"CREATE VIEW a AS SELECT * FROM read_csv_auto('{_B}/01_QueryA_InscricoesPorAno.csv.gz', delim=';', header=true)")
S = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."); D = os.path.join(S, "dadoscreche")
OUT = sys.argv[1]
norm = lambda x: str(x).strip().lstrip('0') or '0'

def canon(b):
    b = (b or '?').upper().strip()
    b = re.sub(r'\s*/\s*.*$', '', b)
    return b.title()

# geo das unidades
wb = openpyxl.load_workbook(os.path.join(D, "OferecimentosEvagas", "Unidades_Unificadas_com_Localizacao.xlsx"), read_only=True)
rows = list(wb['Unidades_Unificadas'].iter_rows(values_only=True)); hdr = [str(x).strip() for x in rows[0]]
U = {}
for r in rows[1:]:
    d = dict(zip(hdr, r))
    try:
        U[norm(d['DESIGNACAO'])] = {'nome': str(d.get('DENOMINACAO') or '').strip(), 'cre': int(float(d['CRE'])),
            'micro': str(d.get('microárea') or ''), 'bairro': str(d.get('BAIRRO') or '').strip(),
            'lat': float(d['LATITUDE']), 'lon': float(d['LONGITUDE'])}
    except Exception: pass

# fila 2025
F = []
for cod, nome, n in c.execute("""SELECT unidade, ANY_VALUE(nome_unidade), COUNT(DISTINCT aluno_anon)
        FROM a WHERE ano=2025 AND situacao='Lista de espera' GROUP BY unidade""").fetchall():
    k = norm(cod)
    if k in U: F.append({'cod': k, 'nome': nome, 'fila': n, **U[k]})

# vagas ociosas da rede parceira (maio/2025)
wb2 = openpyxl.load_workbook(os.path.join(D, "OferecimentosEvagas", "Parceiras2025.xlsx"), read_only=True)
r2 = list(wb2["MAIO -2025"].iter_rows(values_only=True)); grp, h = r2[0], r2[1]
cols = collections.defaultdict(dict); at = None
for i, g in enumerate(grp):
    if g: at = str(g).strip()
    if h[i] and at: cols[at][str(h[i]).strip()] = i
V = []
for r in r2[2:]:
    if not r[1]: continue
    k = norm(r[1])
    if k not in U: continue
    porg = {}
    for g, cs in cols.items():
        if 'Vagas' in cs:
            try:
                x = int(r[cs['Vagas']] or 0)
                if x > 0: porg[g.title()] = x
            except Exception: pass
    if porg:
        try: meta = int(r[4] or 0)
        except Exception: meta = 0
        V.append({'cod': k, 'vagas': sum(porg.values()), 'grupos': porg, 'meta': meta, **U[k]})

def dist(a, b):
    R = 6371; p = math.pi / 180
    dla = (b['lat'] - a['lat']) * p; dlo = (b['lon'] - a['lon']) * p
    x = math.sin(dla / 2) ** 2 + math.cos(a['lat'] * p) * math.cos(b['lat'] * p) * math.sin(dlo / 2) ** 2
    return 2 * R * math.asin(math.sqrt(x))

# vizinhança até 3 km
RAIO = 3.0
for f in F:
    f['viz'] = sorted(
        ({'cod': v['cod'], 'd': round(dist(f, v), 2)} for v in V if dist(f, v) <= RAIO),
        key=lambda z: z['d'])[:6]

Vd = {v['cod']: v for v in V}
comviz = [f for f in F if f['viz']]

# ── agregados de cidade, sem dupla contagem ──
vagas_unicas = {}
for f in comviz:
    for z in f['viz']: vagas_unicas[z['cod']] = Vd[z['cod']]['vagas']

# alocação gulosa: cada vaga usada uma vez só
cap = dict(vagas_unicas); realocaveis = 0
for f in sorted(comviz, key=lambda x: -x['fila']):
    resta = f['fila']
    for z in f['viz']:
        if resta <= 0: break
        usa = min(resta, cap.get(z['cod'], 0))
        cap[z['cod']] -= usa; resta -= usa; realocaveis += usa

# por bairro (dedup das vagas dentro do bairro)
B = collections.defaultdict(lambda: {'fila': 0, 'cre': None, 'vagas': {}, 'un': 0})
for f in comviz:
    b = canon(f['bairro'])
    B[b]['fila'] += f['fila']; B[b]['cre'] = f['cre']; B[b]['un'] += 1
    for z in f['viz']: B[b]['vagas'][z['cod']] = Vd[z['cod']]['vagas']
bairros = sorted(({'nome': b, 'cre': x['cre'], 'fila': x['fila'], 'un': x['un'],
                   'vagas': sum(x['vagas'].values())} for b, x in B.items()),
                 key=lambda r: -r['fila'])

# por CRE
C = collections.defaultdict(lambda: {'fila': 0, 'vagas': {}, 'un': 0})
for f in comviz:
    C[f['cre']]['fila'] += f['fila']; C[f['cre']]['un'] += 1
    for z in f['viz']: C[f['cre']]['vagas'][z['cod']] = Vd[z['cod']]['vagas']
cres = sorted(({'cre': k, 'fila': x['fila'], 'un': x['un'], 'vagas': sum(x['vagas'].values())}
               for k, x in C.items()), key=lambda r: -r['fila'])

json.dump({
    'meta': {'processo': 195, 'ano': 2025, 'raio_km': RAIO, 'ref_vagas': 'maio/2025'},
    'totais': {
        'fila': sum(f['fila'] for f in comviz),
        'vagas': sum(vagas_unicas.values()),
        'realocaveis': realocaveis,
        'unidades': len(comviz),
        'bairros': len(bairros),
    },
    'cres': cres, 'bairros': bairros[:40],
    'focos': [{k: f[k] for k in ('cod', 'nome', 'bairro', 'cre', 'micro', 'fila', 'lat', 'lon', 'viz')} for f in comviz],
    'vagas': {v['cod']: {k: v[k] for k in ('nome', 'bairro', 'cre', 'vagas', 'grupos', 'meta', 'lat', 'lon')} for v in V},
}, open(OUT, 'w'), ensure_ascii=False, separators=(',', ':'))
print(f"ok: {OUT}")
print(f"  fila={sum(f['fila'] for f in comviz)} vagas={sum(vagas_unicas.values())} realocaveis={realocaveis} focos={len(comviz)} bairros={len(bairros)}")
